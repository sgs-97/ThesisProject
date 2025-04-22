import argparse
import os
import re

import pandas as pd
import glob
import openpyxl
from openpyxl import load_workbook


def concatenate_excel_files(directory='.', names=None, output_file='combined.xlsx'):
    """
    Concatenate all Excel files in the current directory into a single Excel file.
    """
    if names is None:
        names = ['summary.xlsx'] # Default names to look for

    # Recursively search all subdirectories for Excel files with names in names
    excel_files = []
    for name in names:
        excel_files.extend(glob.glob(os.path.join(directory, '**', name), recursive=True))

    # Check if there are any Excel files in the current directory
    if excel_files == []:
        print(f"No Excel files found in the {directory} with the specified names: {names}")
        return

    # Create a new workbook for the output
    output_wb = openpyxl.Workbook()
    output_ws = output_wb.active

    for file in excel_files:
        # Load the current workbook
        wb = load_workbook(file)
        ws = wb.active

        # Copy each row from the current workbook to the output workbook
        for row in ws.iter_rows(values_only=False):
            output_ws.append([cell.value for cell in row])


    # Save the output workbook
    output_wb.save(output_file)
    print(f"Combined {len(excel_files)} files into {output_file}")

    return
    # Read and concatenate all Excel files
    dataframes = [pd.read_excel(file) for file in excel_files]
    combined_df = pd.concat(dataframes, ignore_index=True)

    # Write the combined DataFrame to a new Excel file
    combined_df.to_excel(output_file, index=False)

    print(f"Combined {len(excel_files)} files into {output_file}")


if __name__ == '__main__':
    script_name = os.path.basename(__file__)

    parser = argparse.ArgumentParser(description='Description of your script.')
    parser.add_argument('directory', type=str, help='Directory to do the recursive search for xlsx files.')
    parser.add_argument('--names', default=['summary.xlsx'], nargs='+', help='Names of the Excel files to look for. Default: summary.xlsx')
    parser.add_argument('--output_file', default='./combined.xlsx', type=str, help='Output file name for the combined Excel file. Default: ./combined.xlsx')
    parser.add_argument('--verbose', action='store_true', help='Enable verbose mode.')
    args = parser.parse_args()

    verbose = args.verbose
    # Parse other args

    if verbose:
        print(f"[\033[1;33mSTART\033[0m] {script_name}")
        print(f"Directory: {args.directory}")
        print(f"File names: {args.names}")
        print(f"Output file: {args.output_file}")

    if not os.path.exists(args.directory):
        raise FileNotFoundError(f"Directory {args.directory} does not exist.")
    if not os.path.isdir(args.directory):
        raise NotADirectoryError(f"{args.directory} is not a directory.")
    if not os.path.exists(os.path.dirname(args.output_file)):
        os.makedirs(os.path.dirname(args.output_file))
    if not args.output_file.endswith('.xlsx') and not args.output_file.endswith('.xls'):
        print(f"Output file name {args.output_file} does not end with .xlsx or .xls. Fixing...")
        if re.search(r'\..*$', args.output_file):
            args.output_file = re.sub(r'\..*$', '.xlsx', args.output_file)
        else:
            args.output_file = args.output_file + '.xlsx'

    # Call the function to concatenate Excel files
    concatenate_excel_files(directory=args.directory, names=args.names, output_file=args.output_file)

    if args.verbose:
        print(f"[\033[1;32mEXIT\033[0m] {script_name} ended successfully!\033[0m")
